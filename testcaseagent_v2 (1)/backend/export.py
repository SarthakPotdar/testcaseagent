"""
export.py - Exports test cases to XLSX and CSV as bytes (no disk writes for outputs either).
"""

import io
import csv
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLORS = {
    "title_bg": "0F172A", "header_bg": "1E293B", "alt_row": "F8FAFC",
    "white": "FFFFFF", "border": "CBD5E1", "accent": "6366F1",
}
PRI = {
    "High":   {"bg": "FEE2E2", "fg": "991B1B"},
    "Medium": {"bg": "FEF3C7", "fg": "92400E"},
    "Low":    {"bg": "DCFCE7", "fg": "166534"},
}
COLS = [
    ("Test ID", 10), ("Module", 18), ("Title", 35), ("Description", 40),
    ("Preconditions", 28), ("Test Steps", 45), ("Expected Result", 35),
    ("Priority", 12), ("Test Type", 18), ("User Role", 16),
]
KEYS = ["test_id","module","title","description","preconditions",
        "test_steps","expected_result","priority","test_type","user_role"]


def _border():
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)


def create_xlsx_bytes(test_cases: list[dict], summary: dict, doc_summaries: dict = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    t = ws["A1"]
    t.value = f"Test Cases Report  —  {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill = _fill(COLORS["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    for ci, (name, width) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=name)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill(COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 22

    # Rows
    for ri, tc in enumerate(test_cases, 3):
        bg = COLORS["alt_row"] if ri % 2 == 0 else COLORS["white"]
        p = PRI.get(tc.get("priority", "Medium"), PRI["Medium"])
        for ci, key in enumerate(KEYS, 1):
            cell = ws.cell(row=ri, column=ci, value=tc.get(key, ""))
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                horizontal="center" if ci in (1, 8, 9, 10) else "left")
            if ci == 8:
                cell.fill = _fill(p["bg"])
                cell.font = Font(name="Calibri", size=10, bold=True, color=p["fg"])
            elif ci == 1:
                cell.fill = _fill(bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["accent"])
            else:
                cell.fill = _fill(bg)
                cell.font = Font(name="Calibri", size=10, color="1E293B")
        ws.row_dimensions[ri].height = 60

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    for col, w in [("A", 28), ("B", 20), ("C", 28), ("D", 20)]:
        ws2.column_dimensions[col].width = w

    def write_sec(sheet, row, title, data, col_off=0):
        a, b = 1 + col_off, 2 + col_off
        h = sheet.cell(row=row, column=a, value=title)
        h.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        h.fill = _fill(COLORS["header_bg"])
        h.alignment = Alignment(horizontal="center")
        h.border = _border()
        sheet.merge_cells(start_row=row, start_column=a, end_row=row, end_column=b)
        for i, (k, v) in enumerate(data.items(), 1):
            bg = COLORS["alt_row"] if i % 2 == 0 else COLORS["white"]
            for ci, val in [(a, k), (b, v)]:
                c = sheet.cell(row=row+i, column=ci, value=val)
                c.fill = _fill(bg)
                c.font = Font(name="Calibri", size=10, color="1E293B")
                c.border = _border()
                c.alignment = Alignment(horizontal="left" if ci == a else "center")
        return row + len(data) + 2

    ws2.merge_cells("A1:D1")
    t2 = ws2["A1"]
    t2.value = "Test Generation Summary"
    t2.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    t2.fill = _fill(COLORS["title_bg"])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    write_sec(ws2, 3, "Overview", {
        "Total Test Cases": summary.get("total", 0),
        "Documents Processed": len(doc_summaries) if doc_summaries else "N/A",
        "Generated On": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    write_sec(ws2, 3, "By Priority", summary.get("by_priority", {}), col_off=2)
    write_sec(ws2, 9, "By Test Type", summary.get("by_type", {}))
    write_sec(ws2, 9, "By Module", summary.get("by_module", {}), col_off=2)

    if doc_summaries:
        row = 16
        ws2.merge_cells(f"A{row}:D{row}")
        h = ws2[f"A{row}"]
        h.value = "Document Summaries"
        h.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        h.fill = _fill("334155")
        h.alignment = Alignment(horizontal="center")
        row += 1
        for fname, summ in doc_summaries.items():
            ws2.merge_cells(f"A{row}:D{row}")
            fn = ws2[f"A{row}"]
            fn.value = f"{fname}"
            fn.font = Font(name="Calibri", size=10, bold=True, color=COLORS["accent"])
            fn.fill = _fill(COLORS["alt_row"])
            fn.border = _border()
            row += 1
            ws2.merge_cells(f"A{row}:D{row}")
            dc = ws2[f"A{row}"]
            dc.value = summ
            dc.font = Font(name="Calibri", size=10, color="1E293B")
            dc.alignment = Alignment(wrap_text=True, vertical="top")
            dc.fill = _fill(COLORS["white"])
            dc.border = _border()
            ws2.row_dimensions[row].height = 50
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def create_csv_bytes(test_cases: list[dict]) -> bytes:
    buf = io.StringIO()
    headers = [c[0] for c in COLS]
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for tc in test_cases:
        writer.writerow({
            "Test ID": tc.get("test_id",""), "Module": tc.get("module",""),
            "Title": tc.get("title",""), "Description": tc.get("description",""),
            "Preconditions": tc.get("preconditions",""), "Test Steps": tc.get("test_steps",""),
            "Expected Result": tc.get("expected_result",""), "Priority": tc.get("priority",""),
            "Test Type": tc.get("test_type",""), "User Role": tc.get("user_role",""),
        })
    return buf.getvalue().encode("utf-8-sig")
