"""
export.py - Exports test cases to a formatted, professional Excel (.xlsx) file.
"""

import io
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Color palette
COLORS = {
    "header_bg": "1E293B",      # Dark slate
    "header_fg": "FFFFFF",      # White
    "subheader_bg": "334155",   # Slightly lighter slate
    "high_bg": "FEE2E2",        # Light red
    "high_fg": "991B1B",        # Dark red
    "medium_bg": "FEF3C7",      # Light amber
    "medium_fg": "92400E",      # Dark amber
    "low_bg": "DCFCE7",         # Light green
    "low_fg": "166534",         # Dark green
    "alt_row": "F8FAFC",        # Very light gray for alternating rows
    "white": "FFFFFF",
    "border": "CBD5E1",
    "title_bg": "0F172A",       # Very dark for title
    "accent": "6366F1",         # Indigo accent
}

PRIORITY_STYLES = {
    "High":   {"bg": "FEE2E2", "fg": "991B1B"},
    "Medium": {"bg": "FEF3C7", "fg": "92400E"},
    "Low":    {"bg": "DCFCE7", "fg": "166534"},
}

COLUMNS = [
    ("Test ID",        10),
    ("Module",         18),
    ("Title",          35),
    ("Description",    40),
    ("Preconditions",  28),
    ("Test Steps",     45),
    ("Expected Result",35),
    ("Priority",       12),
    ("Test Type",      18),
    ("User Role",      16),
]


def _thin_border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font(size=11, bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=COLORS["header_fg"])


def _cell_font(size=10, bold=False, color="1E293B"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _fill(hex_color: str):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def create_excel_report(test_cases: list[dict], summary: dict, doc_summaries: dict = None) -> bytes:
    """
    Create a polished multi-sheet Excel workbook.

    Returns bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Test Cases ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Test Cases Report  —  Generated {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLORS["header_fg"])
    title_cell.fill = _fill(COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 30

    # Column headers
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = _header_font(size=10)
        cell.fill = _fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, tc in enumerate(test_cases, 3):
        is_alt = (row_idx % 2 == 0)
        row_bg = COLORS["alt_row"] if is_alt else COLORS["white"]
        priority = tc.get("priority", "Medium")
        pri_style = PRIORITY_STYLES.get(priority, PRIORITY_STYLES["Medium"])

        values = [
            tc.get("test_id", ""),
            tc.get("module", ""),
            tc.get("title", ""),
            tc.get("description", ""),
            tc.get("preconditions", ""),
            tc.get("test_steps", ""),
            tc.get("expected_result", ""),
            tc.get("priority", ""),
            tc.get("test_type", ""),
            tc.get("user_role", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col_idx in (1, 8, 9, 10) else "left",
            )

            # Priority cell special styling
            if col_idx == 8:
                cell.fill = _fill(pri_style["bg"])
                cell.font = Font(name="Calibri", size=10, bold=True, color=pri_style["fg"])
            elif col_idx == 1:
                cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["accent"])
                cell.fill = _fill(row_bg)
            else:
                cell.fill = _fill(row_bg)
                cell.font = _cell_font()

        ws.row_dimensions[row_idx].height = 60

    # Freeze panes
    ws.freeze_panes = "A3"

    # Auto filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

    # ── Sheet 2: Summary Dashboard ───────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 20

    def _write_section(sheet, start_row, title, data: dict, col_offset=0):
        a_col = 1 + col_offset
        b_col = 2 + col_offset

        # Section header
        header = sheet.cell(row=start_row, column=a_col, value=title)
        header.font = Font(name="Calibri", size=11, bold=True, color=COLORS["header_fg"])
        header.fill = _fill(COLORS["header_bg"])
        header.alignment = Alignment(horizontal="center")
        header.border = _thin_border()
        sheet.merge_cells(
            start_row=start_row, start_column=a_col,
            end_row=start_row, end_column=b_col
        )

        for i, (k, v) in enumerate(data.items(), 1):
            row = start_row + i
            kc = sheet.cell(row=row, column=a_col, value=k)
            vc = sheet.cell(row=row, column=b_col, value=v)
            bg = COLORS["alt_row"] if i % 2 == 0 else COLORS["white"]
            for c in (kc, vc):
                c.fill = _fill(bg)
                c.font = _cell_font()
                c.border = _thin_border()
                c.alignment = Alignment(horizontal="left" if c == kc else "center")

        return start_row + len(data) + 2

    # Title
    ws2.merge_cells("A1:D1")
    t = ws2["A1"]
    t.value = "📊  Test Generation Summary"
    t.font = Font(name="Calibri", size=16, bold=True, color=COLORS["header_fg"])
    t.fill = _fill(COLORS["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    # Overview
    overview = {
        "Total Test Cases": summary.get("total", 0),
        "Documents Processed": len(doc_summaries) if doc_summaries else "N/A",
        "Generated On": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    _write_section(ws2, 3, "📋  Overview", overview, col_offset=0)

    # Priority breakdown
    _write_section(ws2, 3, "🎯  By Priority", summary.get("by_priority", {}), col_offset=2)

    # Test type breakdown
    _write_section(ws2, 9, "🔬  By Test Type", summary.get("by_type", {}), col_offset=0)

    # Module breakdown
    _write_section(ws2, 9, "📦  By Module", summary.get("by_module", {}), col_offset=2)

    # Document summaries
    if doc_summaries:
        row = 16
        ws2.merge_cells(f"A{row}:D{row}")
        hdr = ws2[f"A{row}"]
        hdr.value = "📄  Document Summaries"
        hdr.font = Font(name="Calibri", size=11, bold=True, color=COLORS["header_fg"])
        hdr.fill = _fill(COLORS["subheader_bg"])
        hdr.alignment = Alignment(horizontal="center")
        row += 1
        for fname, summary_text in doc_summaries.items():
            ws2.merge_cells(f"A{row}:D{row}")
            fn_cell = ws2[f"A{row}"]
            fn_cell.value = f"📎 {fname}"
            fn_cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["accent"])
            fn_cell.fill = _fill(COLORS["alt_row"])
            fn_cell.border = _thin_border()
            row += 1
            ws2.merge_cells(f"A{row}:D{row}")
            desc_cell = ws2[f"A{row}"]
            desc_cell.value = summary_text
            desc_cell.font = _cell_font()
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            desc_cell.fill = _fill(COLORS["white"])
            desc_cell.border = _thin_border()
            ws2.row_dimensions[row].height = 50
            row += 1

    # ── Sheet 3: Raw Graph Data ──────────────────────────────────────────
    # (optional reference sheet)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info(f"Excel report created with {len(test_cases)} test cases.")
    return output.getvalue()


def export_test_cases(
    test_cases: list[dict],
    summary: dict,
    doc_summaries: dict = None,
    output_path: str = None,
) -> bytes:
    """
    Export test cases to Excel.
    If output_path is given, also writes to disk.
    Returns raw bytes of the xlsx.
    """
    xlsx_bytes = create_excel_report(test_cases, summary, doc_summaries)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(xlsx_bytes)
        logger.info(f"Saved to {output_path}")

    return xlsx_bytes
